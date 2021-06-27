# Shopify skips duplicate strings - so need to add more
# Body is needed, so add review title strings too
# Randomize created at
# Add random variations to emails
# 5 Star revies are too much so maybe minimize them a bit

from openpyxl.workbook.workbook import Workbook
from openpyxl import load_workbook
import pandas
import re
import random
import time

titles = [(5, "Good one"), (5, "Totally good and fit"), (5, "fabric is good.. fitting is good"), (5, "Amazing product in perfect price"),
(5, "Good and genuine product"), (5, "It's awesome.."), (5, "Wow I am very happy."), (4, "Nice fitting and cloth"), (4, "Nice product ☺️"),
(4, "Nice t shirts amazing"), (4, "Nice brand"), (4, "Decent fitting"), (4, "Not bad"), (4, "Just go for it 👍😊"),
(4, "Good fabric with reliable price"), (4, "Good👍"), (3, "Not pure cotton"), (3, "Poor quality"), (3, "Size is big.buy a small no."),
(3, "Okay"), (3, "Nothing special"), (2, "Overpriced"), (2, "Not satisfied"), (2, "Colour comes out while washing"),
(2, "Don't waste money"), (2, "Packing is not good"), (2, "bad quality febrics"), (2, "Loose fit."), (2, "Bad quality, colour faded."),
(1, "Poor quality"), (1, "Pooer quality"), (1, "Bad designs"), (1, "I expected more"), (1, "Good"), (1, "Really bad"), (1, "Not buying again")]

reviews = [(5, "Good product, they also delivered faster than I thought."),
(5, "No complaints"), (5, "Decent Quality for the price"), (5, "Product is good"), (5, "Right on the money, good shirts"),
(5, "Love the designs"), (5, "Will buy again, my boyfriend loved this shirt"), (5, "Will buy again, my girlfriend love this shirt"), 
(5, "Ordered for hubby, he like it"), (5, "My brother loved this"), (5, "My sis loved this shirt"), (5, "Really unique designs"),
(5, "Premium quality at cheaper rate, like it"), (5, "Cheaper prices than h&m or zara, you get almost same quality"),
(4, "Good product for the price"), (4, "Really didn't like this design, but over all quality is good"),
(4, "Their customer support is good, made me buy again"), (4, "Good shirts but you can return if you want"),
(4, "My gf loved this shirt"), (4, ""), (4, "My bf like this shirt"), (4, "Good product"), (4, "Nice designs"),
(4, "Ordered second time, nice website"), (4, "I first ordered for my borther, now ordering for my husband, I like this site"),
(4, "Good product but price should have been 100 rs less"), (4, "Will order again"), (4, "Nice shirts"), (4, "I bought 4 shirts like this"),
(3, "Print is nice but little costly"), (3, "This design would look nice in other colors as well, why only white"),
(3, "Delivery was little late but products good"), (3, "Perfect size for my girlfriend"), (3, "Good shirts in this website"),
(3, "I'm reviewing this bcoz they asked me to, I liked the product but delivery was a little late"), (3, "3rd time I'm ordering"),
(3, "Nice place to buy casual shirts"), (3, "Little late delivery"), (2, "Product is costly"), (2, "Didn't like this product"),
(2, "Niec product, slow delivery"), (2, "Slow delivery but good designs"), (2, "I want to buy this design in different color"),
(2, "Didn't like the shirt design"), (2, "Okay is product"), (2, ""), (1, "I generally get deliveries in about 3 days but they took 5 days to deliver"),
(1, "Didn't like the design on this shirt"), (1, "Color faded after around a month"), (1, "Okayish shirt for high price"),
(1, "This is why I prefer buying from a local mall"), (1, "Costly brand"), (1, "My boyfriend didnt like the shirt"), (1, "Not enough plus sizes"),
(1, "Costly for the quality of shirt you get"), (1, "")]

states = ["Andhra Pradesh", "Chhattisgarh", "Haryana", "Gujrat", "Delhi NCR", "Karnataka", "Kerala", "Madhya Pradesh", "Maharashtra", "Punjab", "Rajasthan", "Odisha",
"Tamil Nadu", "Telangana", "Uttar Pradesh", "Uttarakhand", "West Bengal", "Bihar", "Goa"]


def gen_review(handles, num, num_var, ofile, ifile='r_temp_copy.xlsx'):
  wb = load_workbook(filename=ifile)
  sheet = wb.active

  for h in handles:
    r = random.randint(1, 11)
    rating_spread = [5] * 65 + [4] * 15 + [3] * 10 + [2] * r + [1] * (10 - r);

    rev_count = num + random.choice([-1, 1]) * num_var

    for i in range(rev_count):
      rating = random.choice(rating_spread)
      r_texts = list(filter(lambda x: x[0] == rating, reviews))
      rating_title = random.choice(r_texts)[1]
      names = get_names((len(handles) * rev_count) + 100)

      r_name = names[random.randint(0, len(names) - 1)]

      sheet.cell(row=i + 2, column=1, value=h) #handle
      sheet.cell(row=i + 2, column=3, value=rating) #rating
      sheet.cell(row=i + 2, column=4, value=rating_title) #title
      sheet.cell(row=i + 2, column=5, value=f"{r_name[0]} {r_name[1]}") #author
      sheet.cell(row=i + 2, column=6, value=r_name[2]) #email
    
    wb.save(ofile)
    wb.close()


def check_name(name):
  if re.match(r'^[a-zA-Z ]+$', str(name)) and len(str(name).split()) >= 2:
    return True
  else:
    return False

def sanit_names():
  wb = Workbook()
  sheet = wb.active
  sheet.title = 'Sheet 1';

  m_names = pandas.read_csv('Indian-Male-Names.csv');
  f_names = pandas.read_csv('Indian-Female-Names.csv');


  m_valids = list(filter(check_name, list(m_names["name"])))
  f_valids = list(filter(check_name, list(f_names["name"])))

  valids = m_valids + f_valids
  random.shuffle(valids)

  for i, v in enumerate(valids):
    r_sufx = random.choice([random.randint(0, 100), random.choice('abcxyzdefuvw')])
    e_provider = random.choice(["gmail", "yahoo", "outlook", "mail", "quanto", "cxit", "gorrilamail"])

    email = f"{''.join(v.split())}{r_sufx}@{e_provider}.com"
    dat = (v.split()[0].title(), v.split()[1].title(), email)
    for j, x in enumerate(dat):
      sheet.cell(row=i + 1, column=j + 1, value=x)
  
  wb.save("names.xlsx")
  wb.close()

def get_names(count, file="names.xlsx"):
  names = list()
  wb = load_workbook(filename=file)
  sheet = wb.active

  for val in sheet.iter_rows(min_row=2, max_row=count, min_col=1, max_col=3, values_only=True):
    names.append(val)
 
  wb.close()
  return names


def str_time_prop(start, end, time_format, prop):
  stime = time.mktime(time.strptime(start, time_format))
  etime = time.mktime(time.strptime(end, time_format))

  ptime = stime + prop * (etime - stime)

  return time.strftime(time_format, time.localtime(ptime))


def random_date(start, end, prop):
  return str_time_prop(start, end, '%m/%d/%Y %I:%M %p', prop)
    


if __name__ == '__main__':
  print("it runs")
  sanit_names()
  # get_names(2000)
  # gen_review(["abstract-graphic-printed-shirt-navy-blue"], 20, 5, 'output.xlsx')
